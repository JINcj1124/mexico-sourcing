"""输出服务

支持 JSON API 输出、Excel 导出、Webhook 推送。
"""
import json
import os
import logging
from datetime import date, datetime
from pathlib import Path
from typing import List

from backend.config import settings

logger = logging.getLogger(__name__)


class OutputService:
    """输出服务：JSON + Excel 文件生成"""

    def __init__(self):
        self.daily_dir = Path(settings.output_daily_dir)
        self.excel_dir = Path(settings.output_excel_dir)
        self.daily_dir.mkdir(parents=True, exist_ok=True)
        self.excel_dir.mkdir(parents=True, exist_ok=True)

    def to_daily_json(self, products: List[dict], target_date: date,
                      active_festivals: List[dict], stats: dict) -> str:
        """生成每日JSON输出文件"""
        result = {
            "date": target_date.isoformat(),
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "products": products,
            "active_festivals": active_festivals,
            "summary": stats,
        }

        filename = f"{target_date.isoformat()}.json"
        filepath = self.daily_dir / filename

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)

        # 同时写一个 latest.json 方便前端读取
        latest_path = self.daily_dir / "latest.json"
        with open(latest_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)

        logger.info(f"JSON output: {filepath}")
        return str(filepath)

    def to_excel(self, products: List[dict], target_date: date) -> str:
        """生成Excel文件"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = f"选品报告-{target_date.isoformat()}"

            # 表头
            headers = [
                "排名", "评分", "商品名称", "品类", "进货价¥", "空运费¥",
                "打包费¥", "物损费¥", "尾程费¥", "总成本¥", "建议售价¥",
                "建议售价MXN", "溢价倍数", "毛利率", "竞品价MXN", "竞品数量",
                "重量kg", "体积重kg", "店铺", "店铺年限", "发货时效",
                "选品理由", "1688链接"
            ]
            header_fill = PatternFill("solid", fgColor="1F2937")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # 数据行
            for row, p in enumerate(products, 2):
                values = [
                    p.get("rank"), p.get("score"), p.get("product", {}).get("title_zh", ""),
                    p.get("product", {}).get("category", ""),
                    p.get("cost", {}).get("purchase_price_rmb"),
                    p.get("cost", {}).get("air_freight_rmb"),
                    p.get("cost", {}).get("packaging_fee_rmb"),
                    p.get("cost", {}).get("damage_cost_rmb"),
                    p.get("cost", {}).get("last_mile_fee_rmb"),
                    p.get("cost", {}).get("total_cost_rmb"),
                    p.get("pricing", {}).get("suggested_price_rmb"),
                    p.get("pricing", {}).get("suggested_price_mxn"),
                    p.get("pricing", {}).get("markup_ratio"),
                    p.get("pricing", {}).get("estimated_margin"),
                    p.get("competition", {}).get("temu_mx_lowest_mxn") or p.get("competition", {}).get("shopee_mx_lowest_mxn"),
                    p.get("competition", {}).get("competitor_count"),
                    p.get("logistics", {}).get("actual_weight_kg"),
                    p.get("logistics", {}).get("volumetric_weight_kg"),
                    p.get("store", {}).get("name", ""),
                    p.get("store", {}).get("years_active"),
                    p.get("store", {}).get("delivery_hours"),
                    p.get("reason", {}).get("primary", ""),
                    p.get("product", {}).get("source_url", ""),
                ]
                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.border = thin_border
                    if col == 14:  # 毛利率格式
                        cell.number_format = "0.00%"

            # 列宽
            col_widths = [5, 5, 30, 8, 8, 8, 8, 8, 8, 8, 8, 10, 8, 8, 8, 8, 8, 8, 15, 8, 8, 20, 30]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

            filename = f"sourcing_report_{target_date.isoformat()}.xlsx"
            filepath = self.excel_dir / filename
            wb.save(filepath)
            logger.info(f"Excel output: {filepath}")
            return str(filepath)

        except ImportError:
            logger.warning("openpyxl not available, skipping Excel export")
            return ""

    def build_product_dict(self, p, rank: int) -> dict:
        """将内部Product对象转为前端友好的JSON字典"""
        return {
            "id": str(p.id),
            "rank": rank,
            "score": p.score,
            "product": {
                "title_zh": p.title_zh,
                "title_es": p.title_es or "",
                "category": p.category or "",
                "image_url": p.image_url or "",
                "source_url": p.source_url or "",
                "description": p.description or "",
                "description_es": p.description_es or "",
            },
            "cost": {
                "purchase_price_rmb": round(p.purchase_price_rmb, 2),
                "air_freight_rmb": round(p.air_freight_rmb, 2),
                "packaging_fee_rmb": round(settings.packaging_fee, 2),
                "damage_cost_rmb": round(p.purchase_price_rmb * settings.damage_rate, 2),
                "last_mile_fee_rmb": round(settings.last_mile_base_fee, 2),
                "total_cost_rmb": round(p.total_cost_rmb, 2),
            },
            "pricing": {
                "suggested_price_rmb": round(p.suggested_price_rmb, 2),
                "suggested_price_mxn": round(p.suggested_price_mxn, 2),
                "markup_ratio": round(p.markup_ratio, 2),
                "estimated_margin": round(p.estimated_margin, 4),
            },
            "competition": {
                "temu_mx_lowest_mxn": p.temu_mx_lowest_mxn,
                "shopee_mx_lowest_mxn": p.shopee_mx_lowest_mxn,
                "competitor_count": p.competitor_count,
                "price_advantage": "",
            },
            "logistics": {
                "actual_weight_kg": round(p.actual_weight_kg, 3),
                "volumetric_weight_kg": round(p.volumetric_weight_kg, 3),
                "chargeable_weight_kg": round(p.chargeable_weight_kg, 3),
                "dimensions_cm": {
                    "l": round(p.length_cm, 1),
                    "w": round(p.width_cm, 1),
                    "h": round(p.height_cm, 1),
                },
            },
            "reason": {
                "primary": p.reason_primary or "",
                "details": (p.reason_details or "").split("|") if p.reason_details else [],
                "festival_relevance": p.festival_relevance or "",
            },
            "store": {
                "name": p.store_name or "",
                "years_active": p.store_years_active or 0,
                "rating": p.store_rating or 0,
                "delivery_hours": p.store_delivery_hours or 72,
            },
            "tags": p.tags or [],
        }
